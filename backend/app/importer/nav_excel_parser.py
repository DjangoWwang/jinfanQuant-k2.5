"""Manual NAV Excel import parser.

Reads an Excel file containing historical NAV records (typically
provided by fund administrators or operations teams) and returns
a normalised list of dicts.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# Supported column name variations (case-insensitive).
_COLUMN_ALIASES: dict[str, list[str]] = {
    "fund_name": [
        "基金名称", "产品名称", "fund_name", "fund name", "名称",
    ],
    "nav_date": [
        "日期", "净值日期", "估值日期", "nav_date", "date", "nav date",
    ],
    "unit_nav": [
        "单位净值", "unit_nav", "unit nav", "单位净值(元)",
    ],
    "cumulative_nav": [
        "累计净值", "cumulative_nav", "cumulative nav", "累计净值(元)",
        "累计单位净值",
    ],
}


def parse_nav_excel(file_path: str | Path) -> list[dict[str, Any]]:
    """Parse an NAV Excel file into a list of standardised records.

    The function auto-detects column positions by matching header names
    against known aliases, so it works across common format variations
    from different custodians / administrators.

    Args:
        file_path: Path to a ``.xlsx`` file.

    Returns:
        List of dicts, each containing:
            - fund_name (str)
            - nav_date (str): ISO-format date string
            - unit_nav (float)
            - cumulative_nav (float | None)

    Raises:
        ValueError: If required columns cannot be detected.
    """
    file_path = Path(file_path)
    logger.info("Parsing NAV Excel: %s", file_path.name)

    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"Workbook {file_path.name} has no active sheet.")

    # ------- Detect columns -------
    col_map: dict[str, int] = {}
    header_row = 1

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1
    ):
        for cell in row:
            if cell.value is None:
                continue
            cell_text = str(cell.value).strip().lower()
            for canonical, aliases in _COLUMN_ALIASES.items():
                if canonical in col_map:
                    continue
                if cell_text in [a.lower() for a in aliases]:
                    col_map[canonical] = cell.column - 1  # 0-based
                    header_row = row_idx
        # Stop once we've found at least the required columns.
        if "nav_date" in col_map and "unit_nav" in col_map:
            break

    if "nav_date" not in col_map or "unit_nav" not in col_map:
        raise ValueError(
            "Could not detect required columns (nav_date, unit_nav) "
            f"in the first 10 rows of {file_path.name}. "
            f"Detected so far: {list(col_map.keys())}"
        )

    # ------- Read data rows -------
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        nav_date_raw = _cell(row, col_map.get("nav_date"))
        unit_nav_raw = _cell(row, col_map.get("unit_nav"))

        if nav_date_raw is None or unit_nav_raw is None:
            continue

        # Normalise date to ISO string.
        nav_date_str = _normalise_date(nav_date_raw)
        if nav_date_str is None:
            continue

        try:
            unit_nav = float(unit_nav_raw)
        except (ValueError, TypeError):
            continue

        cumulative_nav: float | None = None
        cum_raw = _cell(row, col_map.get("cumulative_nav"))
        if cum_raw is not None:
            try:
                cumulative_nav = float(cum_raw)
            except (ValueError, TypeError):
                pass

        fund_name = str(_cell(row, col_map.get("fund_name")) or "").strip()

        records.append(
            {
                "fund_name": fund_name,
                "nav_date": nav_date_str,
                "unit_nav": unit_nav,
                "cumulative_nav": cumulative_nav,
            }
        )

    wb.close()
    logger.info("Parsed %d NAV records from %s", len(records), file_path.name)
    return records


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def _cell(row: tuple, idx: int | None) -> Any:
    """Safely extract a cell value by 0-based column index."""
    if idx is None or idx >= len(row):
        return None
    cell = row[idx]
    return cell.value if hasattr(cell, "value") else cell


def _normalise_date(value: Any) -> str | None:
    """Convert various date representations to an ISO date string."""
    from datetime import date, datetime

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip().replace("/", "-").replace(".", "-")
        # Quick sanity check: must look like YYYY-MM-DD or similar.
        parts = text.split("-")
        if len(parts) == 3 and len(parts[0]) == 4:
            return text[:10]
    return None
