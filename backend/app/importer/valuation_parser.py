"""Valuation table Excel parser with 4-level hierarchy detection.

Custodian banks provide valuation tables as Excel files.  Each row
contains an ``item_code`` whose length encodes the hierarchy level:

    * 4 characters  -> Level 1 (asset category, e.g. "1102" = stocks)
    * 6 characters  -> Level 2 (sub-category)
    * 8 characters  -> Level 3 (individual security)
    * 14+ characters -> Level 4 (detailed lot / tranche)

This module parses such files into a structured dict ready for
database ingestion, including sub-fund allocations with weights.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)


def _detect_level(item_code: str) -> int:
    """Return the hierarchy level (1-4) based on item_code length."""
    code_len = len(item_code.strip())
    if code_len <= 4:
        return 1
    elif code_len <= 6:
        return 2
    elif code_len <= 8:
        return 3
    else:
        return 4


def _safe_float(value: Any) -> float | None:
    """Convert a cell value to float, returning None on failure."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        # Handle formatted strings like "-18,043.43"
        if isinstance(value, str):
            try:
                return float(value.replace(",", ""))
            except (ValueError, TypeError):
                pass
        return None


# Regex to extract filing number from L4 item codes.
# Pattern: 8-char prefix (e.g. "11090601") + filing_number suffix
_FILING_RE = re.compile(r"^(\d{8})([A-Z0-9]{4,})$", re.IGNORECASE)

# Sub-fund item code prefix for private fund investments
_SUBFUND_PREFIX = "11090601"
# Valuation adjustment prefix
_VALUATION_PREFIX = "11090699"


class ValuationParser:
    """Parse a custodian valuation table Excel file.

    The parser is designed with a strategy pattern in mind: subclass and
    override ``_locate_columns`` / ``_parse_row`` to support different
    custodian formats (e.g. CICC, China Merchants, CITIC).

    Typical usage::

        parser = ValuationParser()
        result = parser.parse("path/to/valuation.xlsx")
        sub_funds = result["sub_fund_allocations"]  # [{filing_number, name, weight, ...}]
    """

    # Column header aliases — matching actual 国信证券 format.
    EXPECTED_COLUMNS: dict[str, list[str]] = {
        "item_code": ["科目代码", "科目编号", "item_code", "code"],
        "item_name": ["科目名称", "item_name", "name"],
        "quantity": ["数量", "份额", "quantity", "shares"],
        "unit_cost": ["单位成本", "unit_cost"],
        "total_cost": ["成本", "成本合计", "total_cost", "cost"],
        "cost_pct": ["成本占净值%", "成本占比", "cost_pct"],
        "market_price": ["市价", "行情价格", "market_price", "price"],
        "market_value": ["市值", "市值合计", "market_value"],
        "mv_pct": ["市值占净值%", "市值占比", "mv_pct", "proportion"],
        "valuation_appreciation": ["估值增值", "浮动盈亏", "appreciation"],
        "halt_info": ["停牌信息", "halt_info"],
    }

    def parse(self, file_path: str | Path) -> dict[str, Any]:
        """Parse the valuation Excel and return structured data.

        Returns:
            A dict with::

                {
                    "file_name": str,
                    "product_name": str | None,
                    "valuation_date": str | None,     # "YYYY-MM-DD"
                    "unit_nav": float | None,
                    "total_nav": float | None,
                    "holdings": [...],                 # all parsed rows
                    "sub_fund_allocations": [          # L4 sub-fund investments
                        {
                            "filing_number": str,      # e.g. "AKE73A"
                            "fund_name": str,
                            "quantity": float,          # shares
                            "unit_cost": float,
                            "cost": float,
                            "cost_weight_pct": float,   # cost / NAV %
                            "market_price": float,      # latest unit NAV
                            "market_value": float,
                            "weight_pct": float,         # market_value / NAV %
                            "appreciation": float,
                        },
                        ...
                    ]
                }
        """
        file_path = Path(file_path)
        logger.info("Parsing valuation table: %s", file_path.name)

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError(f"Workbook {file_path.name} has no active sheet.")

        # Extract metadata from header rows
        product_name = self._extract_product_name(ws)
        valuation_date = self._extract_valuation_date(ws)
        unit_nav = self._extract_unit_nav(ws)

        col_map = self._locate_columns(ws)

        holdings: list[dict[str, Any]] = []
        sub_fund_allocations: list[dict[str, Any]] = []
        total_nav: float | None = None

        header_row = col_map.get("_header_row", 1)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            parsed = self._parse_row(row, col_map)
            if parsed is None:
                continue

            item_code = parsed["item_code"]
            parsed["level"] = _detect_level(item_code)
            holdings.append(parsed)

            # Capture total NAV from summary row
            # In some formats, "资产净值" appears in item_code column
            combined_text = f"{item_code} {parsed.get('item_name', '')}"
            if "资产净值" in combined_text:
                total_nav = parsed.get("market_value")

            # Extract sub-fund allocation (L4 rows under 11090601)
            if item_code.startswith(_SUBFUND_PREFIX) and len(item_code) > 8:
                filing_match = _FILING_RE.match(item_code)
                filing_number = filing_match.group(2) if filing_match else item_code[8:]

                sub_fund_allocations.append({
                    "filing_number": filing_number,
                    "fund_name": parsed["item_name"],
                    "quantity": parsed.get("quantity"),
                    "unit_cost": parsed.get("unit_cost"),
                    "cost": parsed.get("total_cost"),
                    "cost_weight_pct": parsed.get("cost_pct"),
                    "market_price": parsed.get("market_price"),
                    "market_value": parsed.get("market_value"),
                    "weight_pct": parsed.get("mv_pct"),
                    "appreciation": parsed.get("valuation_appreciation"),
                })

        wb.close()

        return {
            "file_name": file_path.name,
            "product_name": product_name,
            "valuation_date": valuation_date,
            "unit_nav": unit_nav,
            "total_nav": total_nav,
            "holdings": holdings,
            "sub_fund_allocations": sub_fund_allocations,
        }

    # ------------------------------------------------------------------
    # Override points for custodian-specific formats
    # ------------------------------------------------------------------

    def _locate_columns(self, ws: Any) -> dict[str, int]:
        """Scan the first few rows to build a column-name -> index mapping."""
        col_map: dict[str, int] = {}
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=10, values_only=False), start=1
        ):
            for cell in row:
                if cell.value is None:
                    continue
                cell_text = str(cell.value).strip()
                for canonical, aliases in self.EXPECTED_COLUMNS.items():
                    if canonical in col_map:
                        continue
                    if cell_text.lower() in [a.lower() for a in aliases]:
                        col_map[canonical] = cell.column - 1  # 0-based
                        col_map["_header_row"] = row_idx
                if len(col_map) > 2:
                    # Found header row — stop searching
                    pass

            # If we found most columns, done
            if len(col_map) >= 6:
                break

        if "_header_row" not in col_map:
            col_map["_header_row"] = 1

        return col_map

    def _parse_row(
        self, row: tuple, col_map: dict[str, int]
    ) -> dict[str, Any] | None:
        """Convert a single worksheet row into a holdings dict."""

        def _cell_value(field: str) -> Any:
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx].value if hasattr(row[idx], "value") else row[idx]

        item_code = _cell_value("item_code")
        if item_code is None or str(item_code).strip() == "":
            return None

        return {
            "item_code": str(item_code).strip(),
            "item_name": str(_cell_value("item_name") or "").strip(),
            "quantity": _safe_float(_cell_value("quantity")),
            "unit_cost": _safe_float(_cell_value("unit_cost")),
            "total_cost": _safe_float(_cell_value("total_cost")),
            "cost_pct": _safe_float(_cell_value("cost_pct")),
            "market_price": _safe_float(_cell_value("market_price")),
            "market_value": _safe_float(_cell_value("market_value")),
            "mv_pct": _safe_float(_cell_value("mv_pct")),
            "valuation_appreciation": _safe_float(
                _cell_value("valuation_appreciation")
            ),
        }

    def _extract_product_name(self, ws: Any) -> str | None:
        """Extract product name from the title area (typically row 2)."""
        for row_idx in range(1, 5):
            for cell in ws[row_idx]:
                if cell.value and "___" in str(cell.value):
                    # Format: "公司___产品名___专用表"
                    parts = str(cell.value).split("___")
                    if len(parts) >= 2:
                        return parts[1].strip()
        return None

    def _extract_valuation_date(self, ws: Any) -> str | None:
        """Extract valuation date from the header area."""
        for row_idx in range(1, 5):
            for cell in ws[row_idx]:
                if cell.value is None:
                    continue
                text = str(cell.value)
                # Pattern: "估值日期：2026-02-27" or "估值日期:2026-02-27"
                m = re.search(r"估值日期[：:]\s*(\d{4}-\d{2}-\d{2})", text)
                if m:
                    return m.group(1)
        return None

    def _extract_unit_nav(self, ws: Any) -> float | None:
        """Extract unit NAV from the header area."""
        for row_idx in range(1, 5):
            for cell in ws[row_idx]:
                if cell.value is None:
                    continue
                text = str(cell.value)
                # Pattern: "单位净值：1.1050"
                m = re.search(r"单位净值[：:]\s*([\d.]+)", text)
                if m:
                    return float(m.group(1))
        return None
