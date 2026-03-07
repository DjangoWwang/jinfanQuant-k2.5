"""Report generation service orchestrating data, attribution, charts, and PDF/Excel."""

from __future__ import annotations

import calendar
import logging
from datetime import date, timedelta
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.engine.attribution import (
    AttributionResult,
    compute_single_period_brinson,
    compute_multi_period_brinson,
    extract_l1_weights,
    get_category_names,
)
from app.engine.metrics import (
    calc_all_metrics,
    calc_interval_metrics,
    calc_return,
    normalize_nav,
)
from app.models.benchmark import (
    CompositeBenchmark,
    CompositeBenchmarkItem,
    IndexNav,
)
from app.models.product import Product, ValuationSnapshot, ValuationItem
from app.services.chart_renderer import (
    render_attribution_bar,
    render_monthly_heatmap,
    render_nav_chart,
    render_pie_chart,
    render_weight_comparison,
)
from app.services.pdf_builder import PDFReportBuilder

logger = logging.getLogger(__name__)


class ReportService:
    """Orchestrates report data gathering, computation, and PDF generation."""

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    async def generate_report(
        self,
        db: AsyncSession,
        product_id: int,
        report_type: str,
        period_start: date,
        period_end: date,
        benchmark_id: int | None = None,
    ) -> bytes:
        """Generate a full PDF report and return its bytes."""
        product = await self._load_product(db, product_id)

        # Build NAV series
        product_nav = await self._build_product_nav(db, product_id)

        # Build benchmark NAV
        bm_id = benchmark_id or product.benchmark_id
        benchmark_nav, benchmark_name = await self._build_benchmark_nav(
            db, bm_id, period_start, period_end
        )

        # Normalize both for chart (rebase to 1.0)
        chart_product_nav = normalize_nav(product_nav) if not product_nav.empty else product_nav
        chart_bench_nav = normalize_nav(benchmark_nav) if benchmark_nav is not None and not benchmark_nav.empty else None

        # Calculate metrics
        metrics = calc_all_metrics(product_nav, start_date=period_start, end_date=period_end)
        interval_metrics = calc_interval_metrics(
            product_nav,
            ["1m", "3m", "6m", "1y", "2y", "3y", "inception"],
            reference_date=period_end,
        )

        # Attribution (only if we have benchmark and snapshots)
        attribution = None
        if bm_id:
            try:
                attribution = await self.compute_attribution(
                    db, product_id, period_start, period_end, bm_id, "monthly"
                )
            except Exception as e:
                logger.warning("Attribution computation failed: %s", e)

        # Get snapshots for holding weights
        curr_snapshot = await self._get_snapshot_near_date(db, product_id, period_end)
        prev_snapshot = await self._get_snapshot_near_date(db, product_id, period_start)

        # Monthly returns
        monthly_returns = self._calc_monthly_returns(product_nav)

        # Format period string
        period_str = f"{period_start.strftime('%Y年%m月')} — {period_end.strftime('%Y年%m月')}"

        # Build PDF
        builder = PDFReportBuilder(
            product_name=product.product_name,
            period=period_str,
            report_type=report_type,
        )

        # Page 1: Header + NAV chart + Metrics
        builder.add_header(
            product_code=product.product_code,
            benchmark_name=benchmark_name,
            custodian=product.custodian,
        )
        builder.add_nav_chart(render_nav_chart(
            chart_product_nav, chart_bench_nav,
            product_name=product.product_name,
            benchmark_name=benchmark_name or "基准",
        ))
        builder.add_metrics_cards(metrics)

        if attribution and attribution.aggregated_categories:
            attr_dicts = [c.model_dump() for c in attribution.aggregated_categories]
            builder.add_attribution_bar_chart(render_attribution_bar(attr_dicts))

        if monthly_returns:
            builder.add_monthly_heatmap(render_monthly_heatmap(monthly_returns))

        if report_type == "monthly":
            # Page 2: Allocation + Weight comparison + Interval table
            builder.add_page_break()

            if curr_snapshot:
                pie_data = self._snapshot_to_pie_data(curr_snapshot)
                if pie_data:
                    builder.add_allocation_pie(render_pie_chart(pie_data))

            if prev_snapshot and curr_snapshot:
                cat_names = get_category_names()
                prev_w = self._snapshot_to_l1_weights(prev_snapshot)
                curr_w = self._snapshot_to_l1_weights(curr_snapshot)
                if prev_w or curr_w:
                    builder.add_weight_comparison(
                        render_weight_comparison(prev_w, curr_w, cat_names)
                    )

            builder.add_interval_metrics_table(interval_metrics)

            # Page 3: Attribution detail table
            if attribution and attribution.aggregated_categories:
                builder.add_page_break()
                builder.add_attribution_table(
                    [c.model_dump() for c in attribution.aggregated_categories]
                )

        return builder.build()

    async def generate_excel_report(
        self,
        db: AsyncSession,
        product_id: int,
        config: dict,
    ) -> bytes:
        """Generate an Excel report and return its bytes.

        Args:
            db: Database session.
            product_id: Product ID.
            config: Dict with report_type, period_start, period_end, benchmark_id.

        Returns:
            Excel file bytes.
        """
        product = await self._load_product(db, product_id)
        product_nav = await self._build_product_nav(db, product_id)

        period_start = config.get("period_start")
        period_end = config.get("period_end")
        if isinstance(period_start, str):
            period_start = date.fromisoformat(period_start)
        if isinstance(period_end, str):
            period_end = date.fromisoformat(period_end)

        bm_id = config.get("benchmark_id") or product.benchmark_id

        # Calculate metrics
        metrics = calc_all_metrics(
            product_nav, start_date=period_start, end_date=period_end
        ) if not product_nav.empty else {}

        # Attribution
        attribution = None
        if bm_id:
            try:
                attribution = await self.compute_attribution(
                    db, product_id, period_start, period_end, bm_id, "monthly"
                )
            except Exception as e:
                logger.warning("Attribution computation failed for Excel: %s", e)

        # Latest holdings
        curr_snapshot = await self._get_snapshot_near_date(db, product_id, period_end)

        # Monthly returns
        monthly_returns = self._calc_monthly_returns(product_nav)

        # ----------------------------------------------------------
        # Build workbook
        # ----------------------------------------------------------
        wb = Workbook()

        # Styles
        header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_font = Font(name="Microsoft YaHei", size=10)
        cell_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        pct_format = "0.00%"
        num_format = "0.000000"

        def _style_header_row(ws, row, col_count):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        def _style_data_cell(ws, row, col, fmt=None, align=None):
            cell = ws.cell(row=row, column=col)
            cell.font = cell_font
            cell.alignment = align or cell_alignment
            cell.border = thin_border
            if fmt:
                cell.number_format = fmt

        # ----------------------------------------------------------
        # Sheet 1: 概览
        # ----------------------------------------------------------
        ws1 = wb.active
        ws1.title = "概览"

        # Product info section
        info_headers = ["项目", "内容"]
        ws1.append(info_headers)
        _style_header_row(ws1, 1, 2)

        info_rows = [
            ("产品名称", product.product_name),
            ("产品代码", product.product_code or "—"),
            ("托管人", product.custodian or "—"),
            ("管理人", product.administrator or "—"),
            ("成立日期", str(product.inception_date) if product.inception_date else "—"),
            ("产品类型", product.product_type or "—"),
            ("报告期间", f"{period_start} 至 {period_end}"),
        ]
        for r_idx, (label, value) in enumerate(info_rows, start=2):
            ws1.cell(row=r_idx, column=1, value=label)
            ws1.cell(row=r_idx, column=2, value=value)
            _style_data_cell(ws1, r_idx, 1, align=left_alignment)
            _style_data_cell(ws1, r_idx, 2, align=left_alignment)

        # Key metrics section (start after a blank row)
        metrics_start = len(info_rows) + 3
        ws1.cell(row=metrics_start, column=1, value="关键指标")
        ws1.cell(row=metrics_start, column=1).font = Font(
            name="Microsoft YaHei", bold=True, size=12, color="1E3A5F"
        )

        metric_headers = ["指标", "数值"]
        ws1.append([])  # blank row
        row_idx = metrics_start + 1
        ws1.cell(row=row_idx, column=1, value=metric_headers[0])
        ws1.cell(row=row_idx, column=2, value=metric_headers[1])
        _style_header_row(ws1, row_idx, 2)

        metric_rows = [
            ("累计收益", metrics.get("total_return"), pct_format),
            ("年化收益", metrics.get("annualized_return"), pct_format),
            ("最大回撤", metrics.get("max_drawdown"), pct_format),
            ("年化波动率", metrics.get("annualized_volatility"), pct_format),
            ("夏普比率", metrics.get("sharpe_ratio"), "0.00"),
            ("卡玛比率", metrics.get("calmar_ratio"), "0.00"),
            ("索提诺比率", metrics.get("sortino_ratio"), "0.00"),
        ]
        for label, value, fmt in metric_rows:
            row_idx += 1
            ws1.cell(row=row_idx, column=1, value=label)
            ws1.cell(row=row_idx, column=2, value=value if value is not None else "—")
            _style_data_cell(ws1, row_idx, 1, align=left_alignment)
            _style_data_cell(ws1, row_idx, 2, fmt=fmt if value is not None else None)

        ws1.column_dimensions["A"].width = 18
        ws1.column_dimensions["B"].width = 25

        # ----------------------------------------------------------
        # Sheet 2: 净值序列
        # ----------------------------------------------------------
        ws2 = wb.create_sheet("净值序列")
        nav_headers = ["日期", "单位净值", "日收益率"]
        ws2.append(nav_headers)
        _style_header_row(ws2, 1, 3)

        if not product_nav.empty:
            daily_returns = product_nav.pct_change()
            for i, (dt, nav_val) in enumerate(product_nav.items()):
                row_idx = i + 2
                ws2.cell(row=row_idx, column=1, value=dt.strftime("%Y-%m-%d"))
                ws2.cell(row=row_idx, column=2, value=float(nav_val))
                ret_val = daily_returns.iloc[i] if i > 0 else None
                ws2.cell(
                    row=row_idx, column=3,
                    value=float(ret_val) if ret_val is not None and pd.notna(ret_val) else None,
                )
                _style_data_cell(ws2, row_idx, 1, align=left_alignment)
                _style_data_cell(ws2, row_idx, 2, fmt=num_format)
                _style_data_cell(ws2, row_idx, 3, fmt=pct_format)

        ws2.column_dimensions["A"].width = 14
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 14

        # ----------------------------------------------------------
        # Sheet 3: 月度收益 (heatmap-style: rows=years, cols=months)
        # ----------------------------------------------------------
        ws3 = wb.create_sheet("月度收益")

        month_labels = [f"{m}月" for m in range(1, 13)]
        ws3.append(["年份"] + month_labels + ["全年"])
        _style_header_row(ws3, 1, 14)

        if monthly_returns:
            # Organize by year -> month
            year_month_map: dict[int, dict[int, float]] = {}
            for entry in monthly_returns:
                yr = entry["year"]
                mo = entry["month"]
                year_month_map.setdefault(yr, {})[mo] = entry["return_pct"]

            # Positive / negative fills for heatmap effect
            pos_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
            neg_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")

            for yr in sorted(year_month_map.keys()):
                row_idx = ws3.max_row + 1
                ws3.cell(row=row_idx, column=1, value=yr)
                _style_data_cell(ws3, row_idx, 1, align=left_alignment)

                annual_return = 1.0
                has_data = False
                for mo in range(1, 13):
                    col = mo + 1
                    val = year_month_map[yr].get(mo)
                    if val is not None:
                        ws3.cell(row=row_idx, column=col, value=val)
                        _style_data_cell(ws3, row_idx, col, fmt=pct_format)
                        # Heatmap coloring
                        cell = ws3.cell(row=row_idx, column=col)
                        cell.fill = pos_fill if val >= 0 else neg_fill
                        annual_return *= (1 + val)
                        has_data = True
                    else:
                        ws3.cell(row=row_idx, column=col, value="")
                        _style_data_cell(ws3, row_idx, col)

                # Annual return in last column
                if has_data:
                    ws3.cell(row=row_idx, column=14, value=annual_return - 1)
                    _style_data_cell(ws3, row_idx, 14, fmt=pct_format)
                else:
                    ws3.cell(row=row_idx, column=14, value="")
                    _style_data_cell(ws3, row_idx, 14)

        ws3.column_dimensions["A"].width = 10
        for col in range(2, 15):
            ws3.column_dimensions[get_column_letter(col)].width = 10

        # ----------------------------------------------------------
        # Sheet 4: 归因分析
        # ----------------------------------------------------------
        ws4 = wb.create_sheet("归因分析")
        attr_headers = [
            "分类", "基准权重", "实际权重", "基准收益",
            "实际收益", "配置效应", "选择效应", "交互效应", "合计",
        ]
        ws4.append(attr_headers)
        _style_header_row(ws4, 1, len(attr_headers))

        if attribution and attribution.aggregated_categories:
            for cat in attribution.aggregated_categories:
                d = cat.model_dump() if hasattr(cat, "model_dump") else cat
                row_idx = ws4.max_row + 1
                ws4.cell(row=row_idx, column=1, value=d.get("category_name", d.get("category", "")))
                _style_data_cell(ws4, row_idx, 1, align=left_alignment)
                values = [
                    d.get("benchmark_weight", 0),
                    d.get("actual_weight", 0),
                    d.get("benchmark_return", 0),
                    d.get("actual_return", 0),
                    d.get("allocation_effect", 0),
                    d.get("selection_effect", 0),
                    d.get("interaction_effect", 0),
                    d.get("total_effect", 0),
                ]
                for c_idx, val in enumerate(values, start=2):
                    ws4.cell(row=row_idx, column=c_idx, value=val)
                    _style_data_cell(ws4, row_idx, c_idx, fmt=pct_format)
        else:
            row_idx = ws4.max_row + 1
            ws4.cell(row=row_idx, column=1, value="暂无归因数据")
            _style_data_cell(ws4, row_idx, 1, align=left_alignment)

        ws4.column_dimensions["A"].width = 18
        for col in range(2, len(attr_headers) + 1):
            ws4.column_dimensions[get_column_letter(col)].width = 14

        # ----------------------------------------------------------
        # Sheet 5: 持仓明细
        # ----------------------------------------------------------
        ws5 = wb.create_sheet("持仓明细")
        holding_headers = [
            "科目代码", "科目名称", "层级", "数量",
            "市值(元)", "占净值比(%)", "备注",
        ]
        ws5.append(holding_headers)
        _style_header_row(ws5, 1, len(holding_headers))

        if curr_snapshot and curr_snapshot.items:
            for item in curr_snapshot.items:
                row_idx = ws5.max_row + 1
                ws5.cell(row=row_idx, column=1, value=item.item_code or "")
                ws5.cell(row=row_idx, column=2, value=item.item_name or "")
                ws5.cell(row=row_idx, column=3, value=item.level)
                ws5.cell(row=row_idx, column=4, value=float(item.quantity) if item.quantity else None)
                ws5.cell(row=row_idx, column=5, value=float(item.market_value) if item.market_value else None)
                ws5.cell(
                    row=row_idx, column=6,
                    value=float(item.value_pct_nav) if item.value_pct_nav else None,
                )
                ws5.cell(row=row_idx, column=7, value=item.remark or "")

                _style_data_cell(ws5, row_idx, 1, align=left_alignment)
                _style_data_cell(ws5, row_idx, 2, align=left_alignment)
                _style_data_cell(ws5, row_idx, 3)
                _style_data_cell(ws5, row_idx, 4, fmt="#,##0.00")
                _style_data_cell(ws5, row_idx, 5, fmt="#,##0.00")
                _style_data_cell(ws5, row_idx, 6, fmt="0.00")
                _style_data_cell(ws5, row_idx, 7, align=left_alignment)
        else:
            row_idx = ws5.max_row + 1
            ws5.cell(row=row_idx, column=1, value="暂无持仓数据")
            _style_data_cell(ws5, row_idx, 1, align=left_alignment)

        ws5.column_dimensions["A"].width = 14
        ws5.column_dimensions["B"].width = 28
        ws5.column_dimensions["C"].width = 8
        ws5.column_dimensions["D"].width = 14
        ws5.column_dimensions["E"].width = 16
        ws5.column_dimensions["F"].width = 14
        ws5.column_dimensions["G"].width = 16

        # ----------------------------------------------------------
        # Save to bytes
        # ----------------------------------------------------------
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def compute_attribution(
        self,
        db: AsyncSession,
        product_id: int,
        period_start: date,
        period_end: date,
        benchmark_id: int | None = None,
        granularity: str = "monthly",
    ) -> AttributionResult:
        """Compute Brinson attribution and return structured result."""
        product = await self._load_product(db, product_id)
        bm_id = benchmark_id or product.benchmark_id

        # Get all snapshots in range
        snapshots = await self._get_snapshots_in_range(db, product_id, period_start, period_end)
        if len(snapshots) < 1:
            return AttributionResult(
                product_id=product_id,
                period_start=period_start,
                period_end=period_end,
                granularity=granularity,
            )

        # Get benchmark weights
        benchmark_weights = await self._get_benchmark_weights(db, bm_id) if bm_id else {}

        # Get benchmark index returns
        benchmark_index_returns = await self._get_benchmark_index_returns(
            db, bm_id, period_start, period_end
        ) if bm_id else {}

        # Build period data for multi-period brinson
        # Split by month boundaries
        periods = self._split_into_periods(period_start, period_end, granularity)
        period_data = []

        cat_names = get_category_names()

        for p_start, p_end in periods:
            # Find snapshot closest to p_end for actual weights
            snap = self._find_closest_snapshot(snapshots, p_end)
            if not snap:
                continue

            actual_w = self._snapshot_to_l1_weights(snap)
            if not actual_w:
                continue

            # Calculate actual category returns from NAV changes
            # For simplicity, use product-level return distributed by weight change
            product_nav = await self._build_product_nav(db, product_id)
            period_return = calc_return(product_nav, p_start, p_end)

            # Estimate per-category actual returns using weight differences
            actual_returns = {}
            for cat in actual_w:
                # Use product return as proxy for each category
                actual_returns[cat] = period_return

            # Benchmark returns per category
            bench_returns = {}
            for cat in set(actual_w) | set(benchmark_weights):
                bench_returns[cat] = benchmark_index_returns.get(cat, 0)

            period_data.append({
                "period_start": p_start,
                "period_end": p_end,
                "actual_weights": actual_w,
                "benchmark_weights": benchmark_weights,
                "actual_returns": actual_returns,
                "benchmark_returns": bench_returns,
                "category_names": cat_names,
            })

        if not period_data:
            return AttributionResult(
                product_id=product_id,
                period_start=period_start,
                period_end=period_end,
                granularity=granularity,
            )

        return compute_multi_period_brinson(period_data, product_id, granularity)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    async def _load_product(self, db: AsyncSession, product_id: int) -> Product:
        result = await db.execute(select(Product).where(Product.id == product_id))
        product = result.scalar_one_or_none()
        if not product:
            raise ValueError(f"Product {product_id} not found")
        return product

    async def _build_product_nav(self, db: AsyncSession, product_id: int) -> pd.Series:
        """Build a pandas Series of product NAV from snapshots."""
        result = await db.execute(
            select(ValuationSnapshot.valuation_date, ValuationSnapshot.unit_nav)
            .where(ValuationSnapshot.product_id == product_id)
            .where(ValuationSnapshot.unit_nav.isnot(None))
            .order_by(ValuationSnapshot.valuation_date)
        )
        rows = result.all()
        if not rows:
            return pd.Series(dtype=float)

        dates = [pd.Timestamp(r.valuation_date) for r in rows]
        navs = [float(r.unit_nav) for r in rows]
        return pd.Series(navs, index=pd.DatetimeIndex(dates), name=f"product_{product_id}")

    async def _build_benchmark_nav(
        self, db: AsyncSession, benchmark_id: int | None,
        start: date, end: date,
    ) -> tuple[pd.Series | None, str | None]:
        """Build benchmark NAV from composite benchmark."""
        if not benchmark_id:
            return None, None

        # Get composite benchmark
        bm_result = await db.execute(
            select(CompositeBenchmark).where(CompositeBenchmark.id == benchmark_id)
        )
        benchmark = bm_result.scalar_one_or_none()
        if not benchmark:
            return None, None

        # Get items
        items_result = await db.execute(
            select(CompositeBenchmarkItem).where(
                CompositeBenchmarkItem.composite_id == benchmark_id
            )
        )
        items = list(items_result.scalars().all())
        if not items:
            return None, benchmark.name

        # Build weighted nav
        nav_series: dict[str, pd.Series] = {}
        weights: dict[str, float] = {}
        for item in items:
            weights[item.index_code] = float(item.weight)
            nav_result = await db.execute(
                select(IndexNav.nav_date, IndexNav.nav_value)
                .where(IndexNav.index_code == item.index_code)
                .where(IndexNav.nav_date >= start)
                .where(IndexNav.nav_date <= end)
                .order_by(IndexNav.nav_date)
            )
            rows = nav_result.all()
            if rows:
                dates = [pd.Timestamp(r.nav_date) for r in rows]
                vals = [float(r.nav_value) for r in rows]
                nav_series[item.index_code] = pd.Series(vals, index=pd.DatetimeIndex(dates))

        if not nav_series:
            return None, benchmark.name

        # Combine: weighted sum of normalized navs
        combined = None
        for code, series in nav_series.items():
            w = weights.get(code, 0)
            normed = series / series.iloc[0] * w
            if combined is None:
                combined = normed
            else:
                combined = combined.add(normed, fill_value=0)

        return combined, benchmark.name

    async def _get_snapshot_near_date(
        self, db: AsyncSession, product_id: int, target: date,
    ) -> ValuationSnapshot | None:
        """Get the snapshot closest to target date (on or before)."""
        result = await db.execute(
            select(ValuationSnapshot)
            .where(ValuationSnapshot.product_id == product_id)
            .where(ValuationSnapshot.valuation_date <= target)
            .order_by(ValuationSnapshot.valuation_date.desc())
            .limit(1)
        )
        return result.scalar_one_or_none()

    async def _get_snapshots_in_range(
        self, db: AsyncSession, product_id: int, start: date, end: date,
    ) -> list[ValuationSnapshot]:
        result = await db.execute(
            select(ValuationSnapshot)
            .where(ValuationSnapshot.product_id == product_id)
            .where(ValuationSnapshot.valuation_date >= start)
            .where(ValuationSnapshot.valuation_date <= end)
            .order_by(ValuationSnapshot.valuation_date)
        )
        return list(result.scalars().all())

    async def _get_benchmark_weights(
        self, db: AsyncSession, benchmark_id: int,
    ) -> dict[str, float]:
        """Get benchmark weights mapped to L1 category codes.

        Returns generic benchmark weights keyed by L1 category codes.
        """
        items_result = await db.execute(
            select(CompositeBenchmarkItem).where(
                CompositeBenchmarkItem.composite_id == benchmark_id
            )
        )
        items = list(items_result.scalars().all())
        # Map index codes to category codes: treat each index as a category
        weights = {}
        for item in items:
            weights[item.index_code] = float(item.weight)
        return weights

    async def _get_benchmark_index_returns(
        self, db: AsyncSession, benchmark_id: int | None,
        start: date, end: date,
    ) -> dict[str, float]:
        """Get per-index returns for the benchmark over the period."""
        if not benchmark_id:
            return {}

        items_result = await db.execute(
            select(CompositeBenchmarkItem).where(
                CompositeBenchmarkItem.composite_id == benchmark_id
            )
        )
        items = list(items_result.scalars().all())

        returns = {}
        for item in items:
            nav_result = await db.execute(
                select(IndexNav.nav_date, IndexNav.nav_value)
                .where(IndexNav.index_code == item.index_code)
                .where(IndexNav.nav_date >= start)
                .where(IndexNav.nav_date <= end)
                .order_by(IndexNav.nav_date)
            )
            rows = nav_result.all()
            if len(rows) >= 2:
                first_val = float(rows[0].nav_value)
                last_val = float(rows[-1].nav_value)
                returns[item.index_code] = (last_val / first_val - 1) if first_val else 0
        return returns

    def _snapshot_to_l1_weights(self, snapshot: ValuationSnapshot) -> dict[str, float]:
        """Extract L1 (level=1) weights from a snapshot's items."""
        items = []
        for item in (snapshot.items or []):
            items.append({
                "item_code": item.item_code or "",
                "level": item.level,
                "value_pct_nav": float(item.value_pct_nav) if item.value_pct_nav else None,
            })
        return extract_l1_weights(items)

    def _snapshot_to_pie_data(self, snapshot: ValuationSnapshot) -> list[dict]:
        """Convert L1 items to pie chart data."""
        cat_names = get_category_names()
        data = []
        for item in (snapshot.items or []):
            if item.level == 1 and item.value_pct_nav and float(item.value_pct_nav) > 0:
                code = item.item_code or ""
                data.append({
                    "name": cat_names.get(code, item.item_name or code),
                    "value": float(item.value_pct_nav),
                })
        return data

    def _calc_monthly_returns(self, nav: pd.Series) -> list[dict]:
        """Calculate monthly returns from NAV series."""
        if nav.empty or len(nav) < 2:
            return []

        monthly = nav.resample("ME").last().dropna()
        rets = monthly.pct_change().dropna()

        result = []
        for dt, ret in rets.items():
            result.append({
                "year": dt.year,
                "month": dt.month,
                "return_pct": float(ret),
            })
        return result

    def _split_into_periods(
        self, start: date, end: date, granularity: str,
    ) -> list[tuple[date, date]]:
        """Split a date range into monthly or weekly sub-periods."""
        periods = []
        current = start

        if granularity == "weekly":
            while current <= end:
                p_end = min(current + timedelta(days=6), end)
                periods.append((current, p_end))
                current = p_end + timedelta(days=1)
        else:  # monthly
            while current <= end:
                year, month = current.year, current.month
                last_day = calendar.monthrange(year, month)[1]
                p_end = min(date(year, month, last_day), end)
                periods.append((current, p_end))
                # Next month
                if month == 12:
                    current = date(year + 1, 1, 1)
                else:
                    current = date(year, month + 1, 1)

        return periods

    @staticmethod
    def _find_closest_snapshot(
        snapshots: list[ValuationSnapshot], target: date,
    ) -> ValuationSnapshot | None:
        """Find snapshot closest to (on or before) the target date."""
        candidates = [s for s in snapshots if s.valuation_date <= target]
        if not candidates:
            return snapshots[0] if snapshots else None
        return max(candidates, key=lambda s: s.valuation_date)


report_service = ReportService()
